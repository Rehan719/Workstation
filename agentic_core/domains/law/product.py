from typing import Dict, Any, List
from agentic_core.omnimedia.factory import OctoOmnimediaGenerator, MultimediaAsset, OutputFormat
from agentic_core.omnimedia.injector import OmnimediaInjector
from agentic_core.constitutional.gaas_validator_v2 import ConstitutionalValidatorV2
from agentic_core.constitutional.ueg_logger import UEGLogger
from agentic_core.constitutional.fallback import FallbackProtocol
from agentic_core.utils.hashing import attach_hash_to_file
import os

class LawProductGenerator(OctoOmnimediaGenerator):
    def __init__(self):
        self.domain = "Law"
        self.validator = ConstitutionalValidatorV2(self.domain)
        self.logger = UEGLogger()
        self.injector = OmnimediaInjector()
        self.fallback = FallbackProtocol(self.domain)

    def generate_infographic(self, data: Dict[str, Any]) -> MultimediaAsset:
        # Precedent timeline or similar
        return MultimediaAsset(
            name=data.get("name", "Law Infographic"),
            asset_type="infographic",
            content=None, # Allow injector to generate mock image
            metadata={"accessibility": {"alt_text": "A timeline of legal precedents."}}
        )

    def generate_video(self, data: Dict[str, Any]) -> MultimediaAsset:
        return MultimediaAsset(
            name=data.get("name", "Law Video"),
            asset_type="video",
            content=None,
            metadata={"accessibility": {"alt_text": "Explainer video about ET1 process."}}
        )

    def generate_audio(self, data: Dict[str, Any]) -> MultimediaAsset:
        return MultimediaAsset(
            name=data.get("name", "Law Audio"),
            asset_type="audio",
            content=None,
            metadata={"accessibility": {"transcript": "This is a recording of the hearing details."}}
        )

    def generate_digital_twin(self, data: Dict[str, Any]) -> MultimediaAsset:
        return MultimediaAsset(
            name=data.get("name", "Courtroom Digital Twin"),
            asset_type="digital_twin",
            content=None,
            metadata={"accessibility": {"alt_text": "3D layout of the Employment Tribunal room."}}
        )

    def generate_et1(self, data: Dict[str, Any]) -> MultimediaAsset:
        """
        Generates an ET1 form as a structured document.
        """
        form_data = data.get("et1_form", {})
        content = f"""
        EMPLOYMENT TRIBUNAL CLAIM (ET1)
        -------------------------------
        Claimant: {form_data.get('claimant_name', 'N/A')}
        Respondent: {form_data.get('respondent_name', 'N/A')}
        Details: {form_data.get('claim_details', 'N/A')}
        """
        return MultimediaAsset("ET1 Form", "document", content)

    def generate_et3(self, data: Dict[str, Any]) -> MultimediaAsset:
        """
        Generates an ET3 response form.
        """
        form_data = data.get("et3_form", {})
        content = f"""
        EMPLOYMENT TRIBUNAL RESPONSE (ET3)
        ----------------------------------
        Respondent: {form_data.get('respondent_name', 'N/A')}
        Defends Claim: {form_data.get('defends_claim', 'Yes')}
        """
        return MultimediaAsset("ET3 Response", "document", content)

    def generate_precedent_timeline(self, data: Dict[str, Any]) -> MultimediaAsset:
        """
        Generates a precedent timeline infographic using matplotlib.
        """
        import matplotlib.pyplot as plt
        events = data.get("precedents", [
            {"date": "2020-01", "event": "Case A"},
            {"date": "2022-06", "event": "Case B"},
            {"date": "2024-03", "event": "Current Case"}
        ])

        dates = [e['date'] for e in events]
        labels = [e['event'] for e in events]

        plt.figure(figsize=(8, 2))
        plt.scatter(dates, [1]*len(dates), c='blue')
        for i, txt in enumerate(labels):
            plt.annotate(txt, (dates[i], 1.05), ha='center')
        plt.yticks([])
        plt.title("Legal Precedent Timeline")

        img_path = "outputs/law_timeline.png"
        os.makedirs("outputs", exist_ok=True)
        try:
            plt.savefig(img_path)
            plt.close()
        except Exception as e:
            # Fallback to logging instead of pass
            self.logger.log_event(self.domain, "GEN_WARNING", {"message": f"Matplotlib failed: {str(e)}"})

        if os.path.exists(img_path):
            with open(img_path, "rb") as f:
                content = f.read()
        else:
            content = b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"

        return MultimediaAsset("Precedent Timeline", "infographic", content)

    def generate_schedule_of_loss_xlsx(self, data: Dict[str, Any]) -> MultimediaAsset:
        """
        Generates a Schedule of Loss Excel workbook with formulas.
        """
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule of Loss"

        ws["A1"] = "Item"
        ws["B1"] = "Amount (£)"
        ws["A2"] = "Basic Award"
        ws["B2"] = 5000
        ws["A3"] = "Compensatory Award"
        ws["B3"] = 15000
        ws["A4"] = "Total"
        ws["B4"] = "=SUM(B2:B3)"

        # Define a named range for the total
        from openpyxl.workbook.defined_name import DefinedName
        new_range = DefinedName("TotalLoss", attr_text="'Schedule of Loss'!$B$4")
        wb.defined_names.add(new_range)

        temp_path = "outputs/schedule_of_loss.xlsx"
        wb.save(temp_path)

        with open(temp_path, "rb") as f:
            content = f.read()
        return MultimediaAsset("Schedule of Loss", "xlsx", content)

    def generate_document(self, data: Dict[str, Any], format: OutputFormat) -> MultimediaAsset:
        return self.generate_et1(data)

    def generate_dashboard(self, data: Dict[str, Any]) -> MultimediaAsset:
        return MultimediaAsset(
            name="Law Dashboard",
            asset_type="dashboard",
            content="<html><body>Dashboard Content</body></html>"
        )

    def create_et1_package(self, claimant_data: Dict[str, Any], target_formats: List[OutputFormat], mode: str = "warning"):
        """
        Pilot: Generates ET1 and Schedule of Loss, validates, and injects into formats.
        """
        # Update validator mode
        self.validator.mode = mode

        # 1. Validation
        validation_result = self.validator.validate_compliance(claimant_data)
        self.logger.log_event(self.domain, "ET1_VALIDATION", {
            "is_valid": validation_result["is_valid"],
            "actual_valid": validation_result.get("actual_valid"),
            "violations": validation_result["violations"]
        })

        # Fallback Check
        fallback_action = self.fallback.evaluate_violations(validation_result["violations"])
        if fallback_action and fallback_action["action"] == "HALT":
            return {"status": "SUSPENDED", "reason": "Constitutional safety halt."}

        if not validation_result["is_valid"] and mode == "reject":
            return {"status": "FAILED", "violations": validation_result["violations"]}

        # 2. Asset Generation
        assets = [
            self.generate_infographic({"name": "Precedent Timeline"}),
            self.generate_digital_twin({"name": "Tribunal Layout"}),
            self.generate_document({"name": "ET1 Form", "text": "Formal ET1 details..."}, OutputFormat.DOCX)
        ]

        # 3. Injection
        results = {}
        for fmt in target_formats:
            target_path = f"ET1_Package_{claimant_data['et1_form']['claimant_name']}.{fmt.value}"
            if fmt == OutputFormat.PDF:
                path = self.injector.inject_into_pdf(target_path, assets)
            elif fmt == OutputFormat.PPTX:
                path = self.injector.inject_into_pptx(target_path, assets)
            elif fmt == OutputFormat.DOCX:
                path = self.injector.inject_into_docx(target_path, assets)
            elif fmt == OutputFormat.XLSX:
                path = self.injector.inject_into_xlsx(target_path, assets)
            elif fmt == OutputFormat.HTML:
                path = self.injector.inject_into_html(target_path, assets)
            else:
                continue

            # 4. Hashing
            with open(path, "rb") as f:
                asset_hash = attach_hash_to_file(path, f.read())

            results[fmt.value] = path
            self.logger.log_event(self.domain, "INJECTION_SUCCESS", {"format": fmt.value, "path": path, "hash": asset_hash})

        return {"status": "SUCCESS", "files": results}
